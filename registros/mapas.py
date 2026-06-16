"""Dominio puro de mapas de corrales: constantes, validacion y parseo de Excel.

No importa modelos de Django para evitar ciclos. openpyxl se importa lazy
dentro de parsear_excel (dependencia solo de desarrollo).
"""

PASILLO_LABEL = "PASILLO"
TORIL_CORRAL_ID = "1"
KINDS_VALIDOS = {"corral", "pasillo", "toril", "pista"}
MAX_GRID = 200


def validar_layout(rows, cols, layout):
	"""Devuelve una lista de mensajes de error. Lista vacia => layout valido."""
	errores = []

	if not isinstance(rows, int) or not isinstance(cols, int) or rows < 1 or cols < 1:
		return ["rows y cols deben ser enteros >= 1."]
	if rows > MAX_GRID or cols > MAX_GRID:
		return [f"rows y cols no pueden superar {MAX_GRID}."]
	if not isinstance(layout, list) or not layout:
		return ["El layout debe ser una lista no vacia de celdas."]

	ocupadas = set()
	labels_corral = []
	toriles = 0
	corrales = 0

	for i, celda in enumerate(layout):
		if not isinstance(celda, dict):
			errores.append(f"Celda {i}: debe ser un objeto.")
			continue

		kind = celda.get("kind")
		if kind not in KINDS_VALIDOS:
			errores.append(f"Celda {i}: kind invalido ({kind!r}).")
			continue

		try:
			row = int(celda["row"])
			col = int(celda["col"])
			row_span = int(celda.get("row_span", 1))
			col_span = int(celda.get("col_span", 1))
		except (KeyError, TypeError, ValueError):
			errores.append(f"Celda {i}: row/col/row_span/col_span deben ser enteros.")
			continue

		if row < 1 or col < 1 or row_span < 1 or col_span < 1:
			errores.append(f"Celda {i}: row/col/spans deben ser >= 1.")
			continue

		if row + row_span - 1 > rows or col + col_span - 1 > cols:
			errores.append(f"Celda {i}: se sale de la grilla {rows}x{cols}.")
			continue

		celda_ocupa = [(r, c) for r in range(row, row + row_span) for c in range(col, col + col_span)]
		solapa = [p for p in celda_ocupa if p in ocupadas]
		if solapa:
			errores.append(f"Celda {i}: se solapa con otra celda en {solapa[0]}.")
		ocupadas.update(celda_ocupa)

		label = str(celda.get("label", "")).strip()
		if kind == "toril":
			toriles += 1
		if kind == "corral":
			corrales += 1
			labels_corral.append(label)

	if corrales == 0:
		errores.append("Debe haber al menos un corral.")
	if toriles > 1:
		errores.append("Solo puede haber un toril.")

	duplicados = sorted({l for l in labels_corral if labels_corral.count(l) > 1})
	if duplicados:
		errores.append(f"Labels de corral duplicados: {duplicados}.")
	if TORIL_CORRAL_ID in labels_corral:
		errores.append(f"Ningun corral puede tener el label reservado '{TORIL_CORRAL_ID}' (TORIL).")

	return errores


def _kind_desde_valor(valor):
	texto = str(valor).strip().upper()
	if texto == "PISTA":
		return "pista", "PISTA"
	if texto.startswith(PASILLO_LABEL):
		return "pasillo", PASILLO_LABEL
	if texto == "TORIL":
		return "toril", "TORIL"
	return "corral", texto


def _agregar_celda(layout, vistas, row, col, row_span, col_span, valor):
	if valor is None or str(valor).strip() == "":
		return
	kind, label = _kind_desde_valor(valor)
	layout.append({
		"row": row,
		"col": col,
		"row_span": row_span,
		"col_span": col_span,
		"kind": kind,
		"label": label,
	})
	for r in range(row, row + row_span):
		for c in range(col, col + col_span):
			vistas.add((r, c))


def parsear_excel(path):
	"""Convierte un Excel (plano dibujado con celdas combinadas) en (rows, cols, layout).

	openpyxl es dependencia solo de desarrollo: se importa aca, no a nivel de modulo.
	"""
	from openpyxl import load_workbook

	wb = load_workbook(path, data_only=True)
	ws = wb.active
	layout = []
	vistas = set()

	for rango in ws.merged_cells.ranges:
		valor = ws.cell(row=rango.min_row, column=rango.min_col).value
		_agregar_celda(
			layout, vistas,
			rango.min_row, rango.min_col,
			rango.max_row - rango.min_row + 1,
			rango.max_col - rango.min_col + 1,
			valor,
		)

	for fila in ws.iter_rows():
		for cell in fila:
			if cell.value is None:
				continue
			if (cell.row, cell.column) in vistas:
				continue
			_agregar_celda(layout, vistas, cell.row, cell.column, 1, 1, cell.value)

	rows = max((c["row"] + c["row_span"] - 1 for c in layout), default=0)
	cols = max((c["col"] + c["col_span"] - 1 for c in layout), default=0)
	return rows, cols, layout
